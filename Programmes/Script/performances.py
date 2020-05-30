# -*- coding: utf-8 -*-
from pathlib import Path    # Appel systeme des paths
import sys  # Gestion de l interpreteur
import openpyxl as xl   # Ecriture dans excel
# Non natif

# Recuperation du chemin du programme
DIRECTORY = Path(__file__).parents[2]
# Ajout du chemin dans la liste des imports de l interpreteur
sys.path.append(str(DIRECTORY))

# Import des scripts necessaires
import Programmes.Script.cloud_of_points as clouds
import Programmes.Class.Timer as Timer
import Programmes.Class.Quickhull as Quickhull
import Programmes.Class.Jarvis as Jarvis
import Programmes.Class.Graham as Graham


NUM_CLOUD = 10000
LIST_NUM = [10, 50, 100, 500, 1000, 5000]  # Taille des nuages

PATH = r"Data.xlsx"


def test() -> None:
    """
    Permet le test des methodes pour calculer les enveloppes convexes
    """
    # Declarations des classes
    G = Graham.Graham()
    J = Jarvis.Jarvis()
    Q = Quickhull.Quickhull()
    Chrono = Timer.Timer()
    classeur = xl.Workbook()
    feuille_G = classeur.create_sheet("Graham", 0)
    feuille_J = classeur.create_sheet("Jarvis", 1)
    feuille_Q = classeur.create_sheet("Quickhull", 2)
    column = 0
    feuille_G.cell(row=2, column=1, value="Nuage n°")
    feuille_J.cell(row=2, column=1, value="Nuage n°")
    feuille_Q.cell(row=2, column=1, value="Nuage n°")
    for num_point in LIST_NUM:
        column += 2
        row = 0
        feuille_G.cell(row=row + 1, column=column,
                       value="Nombre de points: " + str(num_point))
        feuille_G.cell(row=row + 2, column=column, value="Temps:")
        feuille_G.cell(row=row + 2, column=column + 1, value="Itérations:")

        feuille_J.cell(row=row + 1, column=column,
                       value="Nombre de points: " + str(num_point))
        feuille_J.cell(row=row + 2, column=column, value="Temps:")
        feuille_J.cell(row=row + 2, column=column + 1, value="Itérations:")

        feuille_Q.cell(row=row + 1, column=column,
                       value="Nombre de points: " + str(num_point))
        feuille_Q.cell(row=row + 2, column=column, value="Temps:")
        feuille_Q.cell(row=row + 2, column=column + 1, value="Itérations:")

        row = 2
        for i in range(0, NUM_CLOUD):
            row += 1
            cloud = clouds.creation_nuage(
                can_dim=(1000000000, 1000000000), target_point_num=num_point)
            Chrono.start()
            G.convex_hull(cloud)
            Chrono.stop()
            feuille_G.cell(row=row, column=1, value=i + 1)
            feuille_G.cell(row=row, column=column, value=Chrono.time_s)
            feuille_G.cell(row=row, column=column + 1, value=G.itteration)

            Chrono.start()
            J.convex_hull(cloud)
            Chrono.stop()
            feuille_J.cell(row=row, column=1, value=i + 1)
            feuille_J.cell(row=row, column=column, value=Chrono.time_s)
            feuille_J.cell(row=row, column=column + 1, value=J.itteration)

            Chrono.start()
            Q.convex_hull(cloud)
            Chrono.stop()
            feuille_Q.cell(row=row, column=1, value=i + 1)
            feuille_Q.cell(row=row, column=column, value=Chrono.time_s)
            feuille_Q.cell(row=row, column=column + 1, value=Q.itteration)

        classeur.save(PATH)


if __name__ == "__main__":
    test()
    print("FIN")
